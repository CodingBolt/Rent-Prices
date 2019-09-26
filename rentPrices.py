from bs4 import BeautifulSoup
from requests import get
import requests
import re
import math
import openpyxl
import os

'''*****************************************************************************
********************************************************************************
*********************            Over View                 *********************
********************************************************************************
********************************************************************************
This code scrapes data from Daft.ie. It takes a user input of the area(s) to
search and then returns all properties in that area. It saves this data to a
s/sheet of the users descretion. Each area is given a seperate tab named after
the area in question. 

The code then takes user input of how many beds. Using this information it
coallates the price information across all areas searched into a new tab 
called 'Table-Data'. It then calculates the average price for each area by
printing simple excel formulas to the 'Table-Data s/sheet'.

#TODO:
The code needs to be updated to create a bar chart for the average prices 

Contents
Section 1 - Inputs
Section 2 - Procedures
Section 3 - Main Code

********************************************************************************
********************************************************************************
'''

'''*****************************************************************************
********************************************************************************
*********************        Section 1 - Inputs            *********************
********************************************************************************
********************************************************************************
'''

headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
numberOfPropertiesPerPage = 20.0
searchArea = ['dublin-1', 'dublin-2', 'dublin-3','dublin-4', 'dublin-5', 'dublin-6', 'dublin-6w', 'dublin-7', 'dublin-8', 'dublin-9', 'dublin-10', 'dublin-11', 'dublin-12', 'dublin-13', 'dublin-14', 'dublin-15', 'dublin-16', 'dublin-17', 'dublin-18', 'dublin-20', 'dublin-22', 'dublin-24']
filePath = 'rentPrices.xlsx'
bedsPerProperty = [1, 2, 3]

'''*****************************************************************************
********************************************************************************
*********************        Section 2 - Procedures        *********************
********************************************************************************
********************************************************************************
'''

# creates the correct URL for the area to search
def daftURL_Creater(daftURL_part2):
    daftURL_part1 = 'http://daft.ie/dublin-city/residential-property-for-rent/'
    daftURL_part3 = '/?s%5Bignored_agents%5D%5B0%5D=1551'
    daftURL = daftURL_part1 + daftURL_part2 + daftURL_part3
    return daftURL


# prints the data to excel
# open excel
def printToExcel(address, propData, path):
    os.chdir('C:\\Users\\niall\\Desktop\\Delete After\\Daft Price')
    wb = openpyxl.load_workbook(path)

    # create a new tab eg. Dublin-1
    ssData = wb.create_sheet()
    ssData.title = address
    wb.save(path)

    # Set active sheet
    activeSheet = wb[address]
    # Print headers
    activeSheet.cell(row = 1, column = 1).value = 'Address'
    activeSheet.cell(row = 1, column = 2).value = 'Price'
    activeSheet.cell(row = 1, column = 3).value = 'Type'
    activeSheet.cell(row = 1, column = 4).value = 'No Beds'
    activeSheet.cell(row = 1, column = 5).value = 'No Baths'

    # print headers
    rowNum = 2
    # print dictionary data to each column
    for key in propData:
        activeSheet.cell(row = rowNum, column = 1).value = key
        for facility in propData[key]:
            activeSheet.cell(row = rowNum, column = 2).value = propData[key]['price']
            activeSheet.cell(row = rowNum, column = 3).value = propData[key]['type']
            activeSheet.cell(row = rowNum, column = 4).value = propData[key]['beds']
            activeSheet.cell(row = rowNum, column = 5).value = propData[key]['bath']

        rowNum += 1
    wb.save(path)


# Opens excel and collates all prices to 1 tab with headers of the area
def createTable(path, numberOfBeds):
    # get the names of the sheets
    os.chdir('C:\\Users\\niall\\Desktop\\Delete After\\Daft Price')
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
     
    # loop through the list bedsPerProperty
    for beds in range(len(numberOfBeds)):
        # create a new sheet to populate with the prices
        ssSumaryData = wb.create_sheet()
        title = 'Table-Data-' + str(numberOfBeds[beds])
        ssSumaryData.title = title
        wb.save(path)
        # create a dictionary prices{dublin-1:[1000, 1500, 3600, ....]}
        prices = {}
        # loop through sheets to populate the data into prices{}
        for sheet in sheets:
            location = sheet
            activeSheet = wb[sheet]
            if location in ['Sheet', 'Sheet1', 'Table-Data', 'Table-Data-2', 'Table-Data-3']:
                continue
            prices.setdefault(location, [])
            # populate the prices into the list
            for rowNum in range(2, activeSheet.max_row + 1):
                if activeSheet.cell(row = rowNum, column = 4).value == numberOfBeds[beds]:
                    prices[sheet].append(activeSheet.cell(row = rowNum, column = 2).value)
                if activeSheet.cell(row = rowNum, column = 4).value == '':
                    prices[sheet].append(activeSheet.cell(row = rowNum, column = 2).value)


        # print the values of prices to the Table-Data
        colNum = 0
        colList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Z']
        for key in prices:
            colNum += 1
            activeSheet = wb[title]
            # create title
            activeSheet.cell(row = 1, column = colNum).value = key
            # print the formula to calc the avergae price
            activeSheet.cell(row = 2, column = colNum).value = '=IF(ISERROR(AVERAGE(' + colList[colNum - 1] + '4:' + colList[colNum - 1] + '1048576)),0,AVERAGE(' + colList[colNum - 1] + '4:' + colList[colNum - 1] + '1048576))'
            '=AVERAGE(' + colList[colNum - 1] +'4:' + colList[colNum - 1] + '1048576)'
            # print the formula to count the number of properties
            activeSheet.cell(row=3, column = colNum).value = '=COUNT(' + colList[colNum - 1] + '4:' + colList[colNum - 1] + '1048576)'
            # print prices['dublin-1'].value to each row
            for i in range(len(prices[key])):
                activeSheet.cell(row = i + 4, column = colNum).value = prices[key][i]
            
        wb.save(path)


'''*****************************************************************************
********************************************************************************
**********************    Section 3 - Search Daft      *************************
**********************     ------ Main Code -----      *************************
********************************************************************************
********************************************************************************
'''
# Search Daft for all properties in selected areas    
for area in range(len(searchArea)):

    properties = {}
    propertiesCount = 0

    daftURL = daftURL_Creater(searchArea[area])
    # Read in the number of properties at the top of the page
    res = get(daftURL, headers=headers)
    res.raise_for_status()
    daftSoup = BeautifulSoup(res.text, 'lxml')

    # convert the sting to a float
    numberOfPropertiesText = daftSoup.select('.section > strong')[0].get_text()
    numberOfPropertiesRegex = re.compile(r'\d+')
    res = re.search(numberOfPropertiesRegex, numberOfPropertiesText)
    numberOfProperties = float(res.group(0))

    # work out the range for the loop based on 
    # 20 properties per page

    numberOfPages = float(numberOfProperties) / numberOfPropertiesPerPage
    numberOfPages = math.ceil(numberOfPages)
    loopRange = numberOfPages * 20

    # use that for the range of the for loop
    for offset in range(0, loopRange, 20): 
        if offset == 0:
            daftURL = daftURL
            res = get(daftURL, headers=headers)
            res.raise_for_status()
            daftSoup = BeautifulSoup(res.text, 'lxml')
        else:
            daftURLOFF = daftURL + '&offset=' + str(offset) 

            try:
                res = get(daftURLOFF, headers=headers)
                res.raise_for_status()
                daftSoup = BeautifulSoup(res.text, 'lxml')
            except Exception as e:
                break

       
        for elem in daftSoup.select('.box'):
            # clean up the data
            for add in elem.select('h2 > a'):
                #print(add.get_text())
                
                location = add.get_text()
                locRegex = re.compile(r'\s+(- House to Rent|- Apartment to Rent|- Studio apartment to Rent|- Flat to Rent)')
                location = locRegex.sub('', location)
                location = location.strip()

                properties.setdefault(location, {'price': 0, 'type': '', 'beds': '', 'bath': ''})
                # Increase the property count by 1
                propertiesCount += 1

                for cost in elem.select('.price'):
                    fullPriceText = cost.get_text()
                    # create a regex to search for Per week or Per month
                    priceRateRegex = re.compile(r'Per week|Per month')
                    priceRateSearch = re.search(priceRateRegex, fullPriceText)
                    priceRate = priceRateSearch.group(0)

                    # strip out the price amount
                    priceAmountRegex = re.compile(r'\d+,\d+|\d+')
                    priceAmountSearch = re.search(priceAmountRegex, fullPriceText)
                    priceAmountString = priceAmountSearch.group(0)
                    # remove the ',' so that we can convert to int data type
                    priceAmount = re.sub(',', '', priceAmountString)
                    priceAmount = float(priceAmount)    

                    
                    # standardise the priceAmount to monthly
                    if priceRate == 'Per month':
                        priceAmount = priceAmount
                    if priceRate == 'Per week':
                        priceAmount = round(priceAmount * 4.33, 2)

                    # add price to properties
                    properties[location]['price'] = priceAmount
                
            #print(daftSoup.select('.info'))
                for spec in elem.select('.info'):
                    for fac in spec.select('li'):
                        facility = fac.get_text().strip()
                        if 'to Rent' in facility:
                            properties[location]['type'] = facility
                            # Populate the number of beds for a studio apartment
                            if 'Studio apartment to Rent' == facility:
                                properties[location]['beds'] = 1    
                        if 'Beds' in facility or 'Bed' in facility:
                            # create a regex to return just the number
                            noBedsRegex = re.compile(r'\d+')
                            noBedsSearch = re.search(noBedsRegex, facility)
                            noBeds = int(noBedsSearch.group(0))
                            properties[location]['beds'] = noBeds

                        if 'Baths' in facility or 'Bath' in facility:
                            # create a regex to return just the number
                            noBathsRegex = re.compile(r'\d+')
                            noBathsSearch = re.search(noBathsRegex, facility)
                            noBaths = int(noBathsSearch.group(0))
                            properties[location]['bath'] = noBaths

                        #facilities = facilities.strip()
                        #properties.setdefault(facilities, {})

    # Print results to excel
    #print(searchArea[area])
    printToExcel(searchArea[area], properties, filePath)

    # check that the length of properties equals the
    # number of properties on daft
    print('Number of properties: %s \n Properties Count: %a' % (numberOfProperties, propertiesCount))

# create the summary table
createTable(filePath, bedsPerProperty)
